"""
Excel Handler - 엑셀 파일에서 테스트 데이터를 읽어오는 유틸리티
시트별로 데이터를 딕셔너리 리스트로 반환한다.
pandas 1.5.x / Python 3.8.9 호환
"""

import os
import openpyxl
import pandas as pd
from typing import Dict, List, Optional

from config.settings import PATHS


def _read_excel(path, sheet_name):
    # pandas 1.5.x: engine 명시로 openpyxl 사용 (xlsx 기본 엔진)
    return pd.read_excel(
        path,
        sheet_name=sheet_name,
        dtype=str,
        engine="openpyxl",
        keep_default_na=False,  # NaN 자동 변환 방지 (pandas 1.x 호환)
    )


def load_sheet(sheet_name, file_path=None):
    # type: (str, Optional[str]) -> List[Dict]
    """엑셀 파일의 특정 시트를 읽어 딕셔너리 리스트로 반환."""
    path = file_path or PATHS["data"]
    if not os.path.exists(path):
        raise FileNotFoundError("테스트 데이터 파일 없음: {0}".format(path))

    df = _read_excel(path, sheet_name)
    df = df.dropna(how="all")   # 모든 값이 빈 행 제거
    df = df.fillna("")          # 나머지 NaN → 빈 문자열

    # pandas 1.x에서 to_dict 반환값의 숫자형 키 방지: str 변환
    records = df.to_dict(orient="records")
    return [{str(k): v for k, v in row.items()} for row in records]


def load_all_sheets(file_path=None):
    # type: (Optional[str]) -> Dict[str, List[Dict]]
    """모든 시트를 {시트명: 데이터 리스트} 형태로 반환."""
    path = file_path or PATHS["data"]
    if not os.path.exists(path):
        raise FileNotFoundError("테스트 데이터 파일 없음: {0}".format(path))

    with pd.ExcelFile(path, engine="openpyxl") as xl:
        sheet_names = xl.sheet_names  # type: List[str]

    return {sheet: load_sheet(sheet, path) for sheet in sheet_names}


def get_sheet_names(file_path=None):
    # type: (Optional[str]) -> List[str]
    """엑셀 파일의 시트 이름 목록 반환."""
    path = file_path or PATHS["data"]
    if not os.path.exists(path):
        raise FileNotFoundError("테스트 데이터 파일 없음: {0}".format(path))

    with pd.ExcelFile(path, engine="openpyxl") as xl:
        return xl.sheet_names


def create_sample_excel(file_path=None):
    # type: (Optional[str]) -> None
    """테스트용 샘플 엑셀 파일 생성."""
    path = file_path or PATHS["data"]
    dir_path = os.path.dirname(path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)

    wb = openpyxl.Workbook()

    # 시트 1: page_load 테스트
    ws_load = wb.active
    ws_load.title = "tests"
    ws_load.append(["test_id", "test_type", "description"])
    ws_load.append(["TC_LOAD_001", "page_load", "메인 페이지 로딩 확인"])
    ws_load.append(["TC_LOAD_002", "page_load", "타이틀 존재 확인"])

    # 시트 2: click 테스트
    ws_click = wb.create_sheet("click_tests")
    ws_click.append([
        "test_id", "test_type",
        "locator_type", "locator_value", "nth",
        "expected_url_contains", "description",
    ])
    ws_click.append([
        "TC_CLICK_001", "click",
        "partial_link_text", "More information", "",
        "iana.org", "링크 클릭 후 URL 확인",
    ])
    ws_click.append([
        "TC_CLICK_002", "click",
        "class", "btn-item", "2",
        "", "같은 class 2번째 버튼 클릭",
    ])

    # 시트 3: input_and_submit 테스트
    ws_form = wb.create_sheet("form_tests")
    ws_form.append([
        "test_id", "test_type",
        "input_locator_type", "input_locator_value", "input_nth", "input_text",
        "submit_locator_type", "submit_locator_value", "submit_nth",
        "expected_result", "expected_url_contains", "description",
    ])
    ws_form.append([
        "TC_FORM_001", "input_and_submit",
        "css", "input[type='text']", "", "hello",
        "css", "button[type='submit']", "",
        "success", "", "텍스트 입력 후 제출",
    ])
    ws_form.append([
        "TC_FORM_002", "input_and_submit",
        "class", "form-input", "2", "world",
        "class", "btn-ok", "2",
        "success", "", "2번째 입력창 → 2번째 버튼",
    ])

    # 시트 4: 로그인 데이터 (참고용)
    ws_login = wb.create_sheet("login")
    ws_login.append(["test_id", "username", "password", "expected_result", "description"])
    ws_login.append(["TC_LOGIN_001", "admin", "admin123", "success", "정상 로그인"])
    ws_login.append(["TC_LOGIN_002", "admin", "wrongpw",  "fail",    "잘못된 비밀번호"])
    ws_login.append(["TC_LOGIN_003", "",      "admin123", "fail",    "아이디 없음"])

    # 시트 5: 통합 테스트 스위트 ── 모든 타입을 한 시트에서 순서대로 실행
    ws_suite = wb.create_sheet("test_suite")
    SUITE_COLS = [
        "test_id",           # 테스트 케이스 ID
        "test_type",         # page_load | click | click_and_verify |
                             # input | input_and_submit |
                             # verify_text | verify_present | verify_url | select
        "description",       # 테스트 설명
        # ── 클릭 대상 엘리먼트 ────────────────────────────────────────
        "locator_type",      # id | class | css | xpath | name | partial_link_text
        "locator_value",     # 엘리먼트 식별값
        "nth",               # n번째 엘리먼트 (1부터, 비워두면 첫 번째)
        # ── click_and_verify 전용: 클릭 후 확인할 엘리먼트 ────────────
        "verify_locator_type",   # 확인 대상 locator_type
        "verify_locator_value",  # 확인 대상 locator_value
        "verify_nth",            # 확인 대상 nth
        "verify_expect",         # present(기본) | absent | text
        # ── input / input_and_submit 전용 ─────────────────────────────
        "input_text",
        "input_locator_type",
        "input_locator_value",
        "input_nth",
        "submit_locator_type",
        "submit_locator_value",
        "submit_nth",
        # ── verify_text / click_and_verify(text) 전용 ─────────────────
        "expected_text",
        # ── URL 확인 ──────────────────────────────────────────────────
        "expected_url_contains",
        # ── select 전용 ───────────────────────────────────────────────
        "select_by",         # text | value | index
        "select_value",
        # ── 대기 옵션 ─────────────────────────────────────────────────
        "wait_before",       # 액션 전 대기(초)
        "wait_after",        # 클릭 후 검증 전 대기(초) — click_and_verify 전용
        # ── 탐색 제어 ─────────────────────────────────────────────────
        "skip_navigate",     # yes/1/true → 해당 행에서 page.navigate() 생략
                             # 이전 행의 페이지 상태를 유지해 연속 입력·클릭에 사용
    ]
    ws_suite.append(SUITE_COLS)

    # ── 샘플 데이터 (example.com 기준) ──────────────────────────────────────
    E = ""  # 빈 셀 단축
    # skip_navigate 컬럼(마지막)까지 포함: 24개 값
    ws_suite.append([
        "TC_SUITE_001", "page_load", "메인 페이지 로딩 확인",
        E, E, E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, E, E, E, E, E, E,
    ])
    ws_suite.append([
        "TC_SUITE_002", "verify_present", "h1 엘리먼트 존재 확인",
        "tag", "h1", E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, E, E, E, E, E, E,
    ])
    ws_suite.append([
        "TC_SUITE_003", "verify_text", "h1 텍스트 일치 확인",
        "tag", "h1", E,  E, E, E, E,
        E, E, E, E, E, E, E,
        "Example Domain", E, E, E, E, E, E,
    ])
    ws_suite.append([
        "TC_SUITE_004", "click", "More information 링크 클릭",
        "partial_link_text", "More information", E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, "iana.org", E, E, E, E, E,
    ])
    # ── click_and_verify 샘플 ─────────────────────────────────────────────
    ws_suite.append([
        "TC_SUITE_005", "click_and_verify", "버튼 클릭 후 성공 메시지 존재 확인",
        "id", "submit-btn", E,
        "class", "success-msg", E, "present",
        E, E, E, E, E, E, E,
        E, E, E, E, "0.5", E, E,        # wait_after=0.5초
    ])
    ws_suite.append([
        "TC_SUITE_006", "click_and_verify", "삭제 버튼 클릭 후 항목 사라짐 확인",
        "class", "btn-delete", "1",
        "class", "list-item", "1", "absent",
        E, E, E, E, E, E, E,
        E, E, E, E, "1", E, E,          # wait_after=1초
    ])
    ws_suite.append([
        "TC_SUITE_007", "click_and_verify", "탭 클릭 후 탭 콘텐츠 텍스트 확인",
        "class", "tab-menu", "2",
        "css", ".tab-content", E, "text",
        E, E, E, E, E, E, E,
        "탭2 내용", E, E, E, "0.3", E, E,
    ])
    ws_suite.append([
        "TC_SUITE_008", "input_and_submit", "폼 입력 후 제출",
        E, E, E,  E, E, E, E,
        "hello", "css", "input[type='text']", E, "css", "button[type='submit']", E,
        E, E, E, E, E, E, E,
    ])
    ws_suite.append([
        "TC_SUITE_009", "select", "드롭다운 텍스트로 선택",
        "id", "dropdown", E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, E, "text", "옵션1", E, E, E,
    ])
    # ── 로그인 흐름 예시: input(텍스트만) → input(텍스트만) → click(버튼) ──────
    # TC_SUITE_010: 로그인 페이지 접속 (navigate 수행)
    ws_suite.append([
        "TC_SUITE_010", "page_load", "로그인 페이지 접속",
        E, E, E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, E, E, E, E, E, E,   # skip_navigate=빈값 → navigate 실행
    ])
    # TC_SUITE_011: 아이디 입력 (페이지 유지, navigate 생략)
    ws_suite.append([
        "TC_SUITE_011", "input", "아이디 입력",
        "id", "username", E,  E, E, E, E,  # locator_type/locator_value로 입력 필드 지정
        "admin", E, E, E, E, E, E,         # input_text
        E, E, E, E, E, E, "yes",           # skip_navigate=yes → 페이지 유지
    ])
    # TC_SUITE_012: 패스워드 입력 (페이지 유지, navigate 생략)
    ws_suite.append([
        "TC_SUITE_012", "input", "패스워드 입력",
        "id", "password", E,  E, E, E, E,
        "admin123", E, E, E, E, E, E,
        E, E, E, E, E, E, "yes",           # skip_navigate=yes → 페이지 유지
    ])
    # TC_SUITE_013: 로그인 버튼 클릭 (페이지 유지, navigate 생략)
    ws_suite.append([
        "TC_SUITE_013", "click", "로그인 버튼 클릭",
        "id", "login-btn", E,  E, E, E, E,
        E, E, E, E, E, E, E,
        E, E, E, E, E, E, "yes",           # skip_navigate=yes → 페이지 유지
    ])

    wb.save(path)
    print("샘플 엑셀 생성 완료 → {0}".format(path))
